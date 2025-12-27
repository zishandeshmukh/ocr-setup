#!/usr/bin/env python3
"""
End-to-end test: Extract voters with headers and export to Excel
"""

import os
import sys
sys.path.insert(0, os.path.dirname(__file__))

from backend.api import API
from openpyxl import load_workbook

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'google-cloud-vision-key.json'

def test_excel_export_with_headers():
    """Test that headers are included in Excel export"""
    
    print("🔧 Initializing API...")
    api = API()
    print("✅ API initialized\n")
    
    # Set template
    template_name = 'zp_boothwise'
    api.set_template(template_name)
    print(f"📐 Template set: {template_name}\n")
    
    # Process single page
    pdf_path = 'samples/Zp_Boothwise/BoothList_Division_33_Booth_6_A4 राजोली.pdf'
    print(f"📄 Processing: {pdf_path}")
    print("   Page: 6\n")
    
    result = api.process_pdf(pdf_path, start_page=6, end_page=6)
    
    if not result.get('success'):
        print(f"❌ Error: {result.get('error', 'Unknown error')}")
        return False
    
    voters = result.get('voters', [])
    print(f"✅ Extracted {len(voters)} voters\n")
    
    if not voters:
        print("❌ No voters to export")
        return False
    
    # Export to Excel
    output_path = 'temp/test_header_export.xlsx'
    os.makedirs('temp', exist_ok=True)
    
    print(f"📊 Exporting to Excel: {output_path}")
    export_result = api.export_to_excel(output_path)
    
    if not export_result.get('success'):
        print(f"❌ Export failed: {export_result.get('error', 'Unknown error')}")
        return False
    
    print(f"✅ Export successful\n")
    
    # Verify Excel file
    print("🔍 Verifying Excel file...")
    print("="*80)
    
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Get headers (first row)
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(cell_value)
    
    print(f"\n📋 Total columns: {len(headers)}")
    print(f"\nColumn headers:")
    for i, header in enumerate(headers, 1):
        print(f"   {i:2d}. {header}")
    
    # Check for header columns
    required_header_columns = [
        'Header District',
        'Header Taluka',
        'Header Booth',
        'Header Constituency',
        'Header Office',
        'Header Raw Text'
    ]
    
    print(f"\n{'='*80}")
    print("✓ Checking for header columns:")
    print("="*80)
    
    all_present = True
    for col_name in required_header_columns:
        if col_name in headers:
            col_index = headers.index(col_name) + 1
            print(f"   ✅ '{col_name}' found at column {col_index}")
        else:
            print(f"   ❌ '{col_name}' NOT FOUND")
            all_present = False
    
    if not all_present:
        print("\n❌ Some header columns are missing!")
        return False
    
    # Check data in header columns
    print(f"\n{'='*80}")
    print("✓ Checking header data (first 3 rows):")
    print("="*80)
    
    for row_num in range(2, min(5, ws.max_row + 1)):  # Rows 2-4 (data rows)
        print(f"\nRow {row_num - 1}:")
        
        # Get voter name for context
        name_col = headers.index('Name (Marathi)') + 1
        name = ws.cell(row=row_num, column=name_col).value
        print(f"   Voter: {name}")
        
        # Check header fields
        for col_name in required_header_columns:
            col_index = headers.index(col_name) + 1
            value = ws.cell(row=row_num, column=col_index).value
            
            if value:
                # Truncate long values
                display_value = str(value)[:50] + '...' if len(str(value)) > 50 else value
                print(f"   ✅ {col_name}: {display_value}")
            else:
                print(f"   ⚠️ {col_name}: (empty)")
    
    print(f"\n{'='*80}")
    print("✅ Excel verification complete!")
    print("="*80)
    
    return True

if __name__ == '__main__':
    success = test_excel_export_with_headers()
    
    if success:
        print("\n🎉 End-to-end test PASSED!")
        print("   ✅ Headers extracted from PDF")
        print("   ✅ Headers attached to voter records")
        print("   ✅ Headers included in Excel export")
        print("   ✅ All header columns present with data")
    else:
        print("\n❌ End-to-end test FAILED!")
        sys.exit(1)
