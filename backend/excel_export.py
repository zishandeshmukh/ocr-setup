"""
Excel Export Module - Template-Specific
Creates formatted Excel files with voter data based on template type
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# Template-specific column definitions
TEMPLATE_COLUMNS = {
    'boothwise': {
        'headers': [
            'Sr No',
            'Council Name (परिषद नगर)',
            'Ward No (प्रभाग क्रमांक)',
            'Polling Station (मतदान केंद्र)',
            'Part No (यादी भाग क्र)',
            'Polling Address',
            'EPIC',
            'Name (Marathi)',
            'Name (English)',
            'Relation Type',
            'Relation Name (Marathi)',
            'Relation Name (English)',
            'House No',
            'Age',
            'Gender'
        ],
        'data_keys': [
            'serial_excel',  # Will be generated
            'council_name',
            'ward_no',
            'polling_station',
            'part_no',
            'polling_address',
            'epic',
            'name_marathi',
            'name_english',
            'relation_type',
            'relation_name_marathi',
            'relation_name_english',
            'house_no',
            'age',
            'gender'
        ]
    },
    'mahanagpalika': {
        'headers': [
            'Sr No',
            'Corporation (महानगरपालिका)',
            'Ward (प्रभाग)',
            'Part No (यादी भाग क्र)',
            'Address (पत्ता)',
            'EPIC',
            'Name (Marathi)',
            'Name (English)',
            'Relation Type',
            'Relation Name (Marathi)',
            'Relation Name (English)',
            'House No',
            'Age',
            'Gender'
        ],
        'data_keys': [
            'serial_excel',
            'corporation_name',
            'ward',
            'part_no',
            'address',
            'epic',
            'name_marathi',
            'name_english',
            'relation_type',
            'relation_name_marathi',
            'relation_name_english',
            'house_no',
            'age',
            'gender'
        ]
    },
    'wardwise': {
        'headers': [
            'Sr No',
            'Corporation (महानगरपालिका)',
            'Ward (प्रभाग)',
            'Part No (यादी भाग क्र)',
            'Address (पत्ता)',
            'EPIC',
            'Name (Marathi)',
            'Name (English)',
            'Relation Type',
            'Relation Name (Marathi)',
            'Relation Name (English)',
            'House No',
            'Age',
            'Gender'
        ],
        'data_keys': [
            'serial_excel',
            'corporation_name',
            'ward',
            'part_no',
            'address',
            'epic',
            'name_marathi',
            'name_english',
            'relation_type',
            'relation_name_marathi',
            'relation_name_english',
            'house_no',
            'age',
            'gender'
        ]
    },
    'zp_boothwise': {
        'headers': [
            'Sr No',
            'District Council (परिषद जिल्हा)',
            'Election Division (निवार्चन विभाग)',
            'Gan (गण)',
            'Part No (भाग क्र)',
            'Polling Station (मतदान केंद्र)',
            'Address (पत्ता)',
            'EPIC',
            'Name (Marathi)',
            'Name (English)',
            'Relation Type',
            'Relation Name (Marathi)',
            'Relation Name (English)',
            'House No',
            'Age',
            'Gender'
        ],
        'data_keys': [
            'serial_excel',
            'district_council',
            'election_division',
            'gan',
            'part_no',
            'polling_station',
            'address',
            'epic',
            'name_marathi',
            'name_english',
            'relation_type',
            'relation_name_marathi',
            'relation_name_english',
            'house_no',
            'age',
            'gender'
        ]
    },
    'boothlist_division': {
        'headers': [
            'Sr No',
            'District Council (जिल्हा परिषद)',
            'Election Division (निवडणूक विभाग)',
            'Gan (गण)',
            'Part No (भाग क्र)',
            'Polling Station (मतदान केंद्र)',
            'Address (पत्ता)',
            'EPIC',
            'Name (Marathi)',
            'Name (English)',
            'Relation Type',
            'Relation Name (Marathi)',
            'Relation Name (English)',
            'House No',
            'Age',
            'Gender'
        ],
        'data_keys': [
            'serial_excel',
            'district_council',
            'election_division',
            'gan',
            'part_no',
            'polling_station',
            'address',
            'epic',
            'name_marathi',
            'name_english',
            'relation_type',
            'relation_name_marathi',
            'relation_name_english',
            'house_no',
            'age',
            'gender'
        ]
    },
    'ac_wise_low_quality': {
        'headers': [
            'Sr No',
            'Assembly Constituency (विधानसभा मतदारसंघ)',
            'Division (विभाग)',
            'Part No (यादी भाग क्रमांक)',
            'EPIC',
            'Name (Marathi)',
            'Name (English)',
            'Relation Type',
            'Relation Name (Marathi)',
            'Relation Name (English)',
            'House No',
            'Age',
            'Gender'
        ],
        'data_keys': [
            'serial_excel',
            'assembly_constituency',
            'division',
            'part_no',
            'epic',
            'name_marathi',
            'name_english',
            'relation_type',
            'relation_name_marathi',
            'relation_name_english',
            'house_no',
            'age',
            'gender'
        ]
    }
}

# Add alias for mahanagarpalika (with 'nagar') to ensure both spellings work
TEMPLATE_COLUMNS['mahanagarpalika'] = TEMPLATE_COLUMNS['mahanagpalika']

# Add aliases for different template key versions
TEMPLATE_COLUMNS['ward_wise_data'] = TEMPLATE_COLUMNS['wardwise']
TEMPLATE_COLUMNS['mahanagpalika_data'] = TEMPLATE_COLUMNS['mahanagpalika']

# Default generic columns (for templates not yet customized)
DEFAULT_HEADERS = [
    'Page Number',
    'Assembly Name',
    'Part No',
    'Polling Station',
    'Polling Address',
    'Serial No',
    'EPIC',
    'Name (Marathi)',
    'Name (English)',
    'Relation Type',
    'Relation Name (Marathi)',
    'Relation Name (English)',
    'House No',
    'Age',
    'Gender',
    'Header Raw Text'
]


def parse_boothwise_header(raw_header):
    """
    Parse boothwise raw header text into structured fields
    
    Sample:
    मतदान केंद्र : १ पेपर मिल मंगल कार्यालय उत्तरेकडील भाग खोली
    परिषद नगर बल्लारपूर
    प्रभाग क्र : १ - प्रभाग क्र . १
    यादी भाग क्र . १६२ : ४ - बिहारी किराणा जवळील परिसर गोकुल
    """
    result = {
        'council_name': '',
        'ward_no': '',
        'polling_station': '',
        'part_no': '',
        'polling_address': ''
    }
    
    if not raw_header:
        return result
    
    # Parse council name (परिषद नगर X)
    council_match = re.search(r'परिषद\s*नगर\s*([^\n]+)', raw_header)
    if council_match:
        result['council_name'] = council_match.group(1).strip()
    
    # Parse ward number (प्रभाग क्र : X)
    ward_match = re.search(r'प्रभाग\s*क्र\s*[:\s]*([^\n]+)', raw_header)
    if ward_match:
        result['ward_no'] = ward_match.group(1).strip()
    
    # Parse polling station (मतदान केंद्र : X) - capture everything after मतदान केंद्र
    station_match = re.search(r'मतदान\s*केंद्र\s*[:\s]*(.+?)(?:\n|$)', raw_header)
    if station_match:
        result['polling_station'] = station_match.group(1).strip()
    
    # Parse part number (यादी भाग क्र . X)
    part_match = re.search(r'यादी\s*भाग\s*क्र[.\s:]*(\d+|[०-९]+)', raw_header)
    if part_match:
        result['part_no'] = part_match.group(1).strip()
    
    # Parse address - text after the second : in यादी भाग line
    # Pattern: यादी भाग क्र . १६२ : ४ - [address]
    addr_match = re.search(r'यादी\s*भाग\s*क्र[^:]*:\s*[^\-–]*[-–]\s*(.+?)(?:\n|$)', raw_header)
    if addr_match:
        result['polling_address'] = addr_match.group(1).strip()
    
    return result


def parse_mahanagarpalika_header(raw_header):
    """
    Parse Mahanagarpalika raw header text into structured fields
    
    Sample input:
    महानगरपालिका चंद्रपूर
    भानापेठ ११ – प्रभाग क्र : -
    यादी भाग क्र . १५८ : १ - जटपुरागेटरामाला मार्ग किल्ला लगत् चंद्रपुर
    """
    result = {
        'corporation_name': '',
        'ward': '',
        'part_no': '',
        'address': ''
    }
    
    if not raw_header:
        return result
    
    # Parse corporation name (महानगरपालिका X or चंद्रपूर महानगरपालिका)
    corp_match = re.search(r'महानगरपालिका\s*([^\n]+)', raw_header)
    if corp_match:
        result['corporation_name'] = 'महानगरपालिका ' + corp_match.group(1).strip()
    else:
        # Alternative: look for "चंद्रपूर महानगरपालिका" anywhere
        corp_alt = re.search(r'(चंद्रपूर\s*महानगरपालिका)', raw_header)
        if corp_alt:
            result['corporation_name'] = corp_alt.group(1).strip()
        else:
            # Default: if header has प्रभाग क्र, likely a mahanagarpalika
            if 'प्रभाग' in raw_header:
                result['corporation_name'] = 'महानगरपालिका'
    
    # Parse ward (भानापेठ ११ – प्रभाग क्र)
    ward_match = re.search(r'([^\n]*प्रभाग\s*क्र[^\n]*)', raw_header)
    if ward_match:
        result['ward'] = ward_match.group(1).strip()
    
    # Parse part number (यादी भाग क्र . १५८)
    part_match = re.search(r'यादी\s*भाग\s*क्र[.\s:]*(\d+|[०-९]+)', raw_header)
    if part_match:
        result['part_no'] = part_match.group(1).strip()
    
    # Parse address (after : १ - in यादी भाग line)
    # Pattern: यादी भाग क्र . १५८ : १ - [address]
    addr_match = re.search(r'यादी\s*भाग\s*क्र[^:]*:\s*\d*\s*[-–]\s*([^\n]+)', raw_header)
    if addr_match:
        result['address'] = addr_match.group(1).strip()
    
    return result


def parse_zp_boothwise_header(raw_header):
    """
    Parse ZP Boothwise raw header text into structured fields
    
    Sample input:
    परिषद जिल्हा चंद्रपुर
    मारोडा - निवार्चन निवडणूक विभाग : राजोली - गण ३३
    कोळसा : १ - भाग क्र . ६ यादी
    कोळसा नविन : १ मतदान केंद्र कोळसा , जि.प.प्रा.शाळा पत्ता :
    """
    result = {
        'district_council': '',
        'election_division': '',
        'gan': '',
        'part_no': '',
        'polling_station': '',
        'address': ''
    }
    
    if not raw_header:
        return result
    
    # Parse district council - first line containing "जिल्हा" or "परिषद"
    district_match = re.search(r'(परिषद[^\n]*जिल्हा[^\n]*|जिल्हा[^\n]*परिषद[^\n]*)', raw_header)
    if district_match:
        result['district_council'] = district_match.group(1).strip()
    
    # Parse election division - line containing निवार्चन or विभाग
    # Get everything from start of line up to गण
    division_match = re.search(r'([^\n]*(?:निवार्चन|विभाग)[^\n]*?)(?:\s*[-–]\s*गण|\s*गण)', raw_header)
    if division_match:
        result['election_division'] = division_match.group(1).strip()
    else:
        # Fallback - just get the line with विभाग
        division_match2 = re.search(r'([^\n]*विभाग[^\n]+)', raw_header)
        if division_match2:
            result['election_division'] = division_match2.group(1).strip()
    
    # Parse Gan (गण X)
    gan_match = re.search(r'गण\s*[:\s]*(\d+|[०-९]+)', raw_header)
    if gan_match:
        result['gan'] = gan_match.group(1).strip()
    
    # Parse part number (भाग क्र . X or यादी भाग X)
    part_match = re.search(r'भाग\s*क्र[.\s:]*(\d+|[०-९]+)', raw_header)
    if part_match:
        result['part_no'] = part_match.group(1).strip()
    
    # Parse polling station - line containing मतदान केंद्र
    station_match = re.search(r'मतदान\s*केंद्र[:\s]*([^,\n]+)', raw_header)
    if station_match:
        result['polling_station'] = station_match.group(1).strip()
    
    # Parse address - text between comma and पत्ता (e.g., जि.प.प्रा.शाळा)
    # Pattern: ...मतदान केंद्र कोळसा , जि.प.प्रा.शाळा पत्ता :
    addr_match = re.search(r',\s*([^,\n]+?)\s*पत्ता', raw_header)
    if addr_match:
        result['address'] = addr_match.group(1).strip()
    else:
        # Fallback - try to get anything after पत्ता :
        addr_match2 = re.search(r'पत्ता\s*[:\s]+([^\n]+)', raw_header)
        if addr_match2 and addr_match2.group(1).strip():
            result['address'] = addr_match2.group(1).strip()
    
    return result


def parse_ac_wise_header(raw_header):
    """
    Parse AC Wise Low Quality raw header text into structured fields
    
    Sample input:
    विधानसभा मतदारसंघ क्रमांक आणि नाव : 72-बल्लारपूर
    विभाग क्रमांक आणि नाव 1-पायली भटाळी
    """
    result = {
        'assembly_constituency': '',
        'division': ''
    }
    
    if not raw_header:
        return result
    
    
    # Parse assembly constituency (विधानसभा मतदारसंघ क्रमांक आणि नाव : X)
    assembly_match = re.search(r'विधानसभा\s*मतदारसंघ\s*क्रमांक\s*आणि\s*नाव\s*[:\s]*([^\n]+)', raw_header)
    if assembly_match:
        result['assembly_constituency'] = assembly_match.group(1).strip()
    
    # Parse division (विभाग क्रमांक आणि नाव X)
    division_match = re.search(r'विभाग\s*क्रमांक\s*आणि\s*नाव\s*[:\s]*([^\n]+)', raw_header)
    if division_match:
        result['division'] = division_match.group(1).strip()
        
    # Parse Part No (यादी भाग क्रमांक : X)
    part_match = re.search(r'यादी\s*भाग\s*क्रमांक\s*[:\s]*(\d+)', raw_header)
    if part_match:
        result['part_no'] = part_match.group(1).strip()
    
    return result


def export_to_excel(voters, output_path, template='default'):
    """
    Export voters to formatted Excel file
    
    Args:
        voters: List of voter dictionaries
        output_path: Path to save Excel file
        template: Template type (boothwise, ac_wise, etc.)
    """
    if not voters or len(voters) == 0:
        raise ValueError("Cannot export: No voter records provided")
    
    print(f"📊 Excel Export: {len(voters)} records, Template: {template}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Voter Data"
    
    # Get template-specific columns or use default
    
    if template_config:
        headers = template_config['headers']
        data_keys = template_config['data_keys']
    else:
        headers = DEFAULT_HEADERS
        data_keys = None  # Use default mapping
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sort voters by extraction order (matches UI)
    def get_sort_key(v):
        page = v.get('page_number', 0)
        ext_order = v.get('extraction_order', 99999)
        return (page, ext_order)
    
    sorted_voters = sorted(voters, key=get_sort_key)
    
    # Process voters based on template
    if template_key == 'boothwise':
        for row_num, voter in enumerate(sorted_voters, 2):
            # Parse header into structured fields
            raw_header = voter.get('header_raw_text', '')
            parsed = parse_boothwise_header(raw_header)
            
            # Merge parsed header with existing voter data
            voter_data = {
                'serial_excel': row_num - 1,
                'council_name': parsed.get('council_name') or voter.get('header_booth', ''),
                'ward_no': parsed.get('ward_no', ''),
                'polling_station': parsed.get('polling_station') or voter.get('polling_station', ''),
                'part_no': parsed.get('part_no') or voter.get('part_no', ''),
                'polling_address': parsed.get('polling_address') or voter.get('polling_address', ''),
                'epic': voter.get('epic', ''),
                'name_marathi': voter.get('name_marathi', ''),
                'name_english': voter.get('name_english', ''),
                'relation_type': voter.get('relation_type', ''),
                'relation_name_marathi': voter.get('relation_name_marathi', ''),
                'relation_name_english': voter.get('relation_name_english', ''),
                'house_no': voter.get('house_no', ''),
                'age': voter.get('age', ''),
                'gender': voter.get('gender', '')
            }
            
            # Write data
            for col_num, key in enumerate(data_keys, 1):
                ws.cell(row=row_num, column=col_num, value=voter_data.get(key, ''))
    elif template_key in ('mahanagpalika', 'mahanagarpalika', 'wardwise', 'ward_wise_data'):
        for row_num, voter in enumerate(sorted_voters, 2):
            # Parse header into structured fields
            raw_header = voter.get('header_raw_text', '')
            parsed = parse_mahanagarpalika_header(raw_header)
            
            # Merge parsed header with existing voter data
            voter_data = {
                'serial_excel': row_num - 1,
                'corporation_name': parsed.get('corporation_name', ''),
                'ward': parsed.get('ward', ''),
                'part_no': parsed.get('part_no') or voter.get('part_no', ''),
                'address': parsed.get('address') or voter.get('polling_address', ''),
                'epic': voter.get('epic', ''),
                'name_marathi': voter.get('name_marathi', ''),
                'name_english': voter.get('name_english', ''),
                'relation_type': voter.get('relation_type', ''),
                'relation_name_marathi': voter.get('relation_name_marathi', ''),
                'relation_name_english': voter.get('relation_name_english', ''),
                'house_no': voter.get('house_no', ''),
                'age': voter.get('age', ''),
                'gender': voter.get('gender', '')
            }
            
            # Write data
            for col_num, key in enumerate(data_keys, 1):
                ws.cell(row=row_num, column=col_num, value=voter_data.get(key, ''))
    elif template_key in ('zp_boothwise', 'boothlist_division'):
        for row_num, voter in enumerate(sorted_voters, 2):
            # Parse header into structured fields
            raw_header = voter.get('header_raw_text', '')
            parsed = parse_zp_boothwise_header(raw_header)
            
            # Merge parsed header with existing voter data
            voter_data = {
                'serial_excel': row_num - 1,
                'district_council': parsed.get('district_council', ''),
                'election_division': parsed.get('election_division', ''),
                'gan': parsed.get('gan', ''),
                'part_no': parsed.get('part_no') or voter.get('part_no', ''),
                'polling_station': parsed.get('polling_station') or voter.get('polling_station', ''),
                'address': parsed.get('address') or voter.get('polling_address', ''),
                'epic': voter.get('epic', ''),
                'name_marathi': voter.get('name_marathi', ''),
                'name_english': voter.get('name_english', ''),
                'relation_type': voter.get('relation_type', ''),
                'relation_name_marathi': voter.get('relation_name_marathi', ''),
                'relation_name_english': voter.get('relation_name_english', ''),
                'house_no': voter.get('house_no', ''),
                'age': voter.get('age', ''),
                'gender': voter.get('gender', '')
            }
            
            # Write data
            for col_num, key in enumerate(data_keys, 1):
                ws.cell(row=row_num, column=col_num, value=voter_data.get(key, ''))
    elif template_key == 'ac_wise_low_quality':
        for row_num, voter in enumerate(sorted_voters, 2):
            # Parse header into structured fields
            raw_header = voter.get('header_raw_text', '')
            parsed = parse_ac_wise_header(raw_header)
            
            # Merge parsed header with existing voter data
            voter_data = {
                'serial_excel': row_num - 1,
                'assembly_constituency': parsed.get('assembly_constituency', ''),
                'division': parsed.get('division', ''),
                'part_no': parsed.get('part_no', ''),
                'epic': voter.get('epic', ''),
                'name_marathi': voter.get('name_marathi', ''),
                'name_english': voter.get('name_english', ''),
                'relation_type': voter.get('relation_type', ''),
                'relation_name_marathi': voter.get('relation_name_marathi', ''),
                'relation_name_english': voter.get('relation_name_english', ''),
                'house_no': voter.get('house_no', ''),
                'age': voter.get('age', ''),
                'gender': voter.get('gender', '')
            }
            
            # Write data
            for col_num, key in enumerate(data_keys, 1):
                ws.cell(row=row_num, column=col_num, value=voter_data.get(key, ''))
    else:
        # Default export (for other templates)
        for row_num, voter in enumerate(sorted_voters, 2):
            excel_serial = row_num - 1
            ws.cell(row=row_num, column=1, value=voter.get('page_number', ''))
            ws.cell(row=row_num, column=2, value=voter.get('assembly_name', ''))
            ws.cell(row=row_num, column=3, value=voter.get('part_no', ''))
            ws.cell(row=row_num, column=4, value=voter.get('polling_station', ''))
            ws.cell(row=row_num, column=5, value=voter.get('polling_address', ''))
            ws.cell(row=row_num, column=6, value=excel_serial)
            ws.cell(row=row_num, column=7, value=voter.get('epic', ''))
            ws.cell(row=row_num, column=8, value=voter.get('name_marathi', ''))
            ws.cell(row=row_num, column=9, value=voter.get('name_english', ''))
            ws.cell(row=row_num, column=10, value=voter.get('relation_type', ''))
            ws.cell(row=row_num, column=11, value=voter.get('relation_name_marathi', ''))
            ws.cell(row=row_num, column=12, value=voter.get('relation_name_english', ''))
            ws.cell(row=row_num, column=13, value=voter.get('house_no', ''))
            ws.cell(row=row_num, column=14, value=voter.get('age', ''))
            ws.cell(row=row_num, column=15, value=voter.get('gender', ''))
            ws.cell(row=row_num, column=16, value=voter.get('header_raw_text', ''))
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add filters
    ws.auto_filter.ref = ws.dimensions
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel file saved: {output_path}")
    print(f"✅ Exported {len(sorted_voters)} records")
